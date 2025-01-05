from collections.abc import Generator
from typing import Any
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

def get_cell_styles(cell) -> dict:
    """セルのすべてのスタイル情報を取得"""
    styles = {}
    
    # セル結合情報の取得
    if cell.parent.merged_cells:
        for merged_range in cell.parent.merged_cells.ranges:
            if cell.coordinate in merged_range:
                styles['merge'] = {
                    'start': merged_range.coord.split(':')[0],
                    'end': merged_range.coord.split(':')[1]
                }
                break
    
    # 罫線情報（すべての罫線情報を取得）
    if cell.border:
        borders = {}
        for side in ['top', 'bottom', 'left', 'right']:
            border = getattr(cell.border, side)
            if border and border.style:
                borders[side] = border.style
        if borders:
            styles['borders'] = borders
    
    # 背景色（パターンタイプと色情報を取得）
    if cell.fill:
        if hasattr(cell.fill, 'patternType') and cell.fill.patternType:
            if cell.fill.patternType == 'solid':
                if (cell.fill.start_color and 
                    cell.fill.start_color.rgb and 
                    isinstance(cell.fill.start_color.rgb, str)):
                    # 透明の場合は除外
                    if cell.fill.start_color.rgb != '00000000':
                        styles['background'] = cell.fill.start_color.rgb
    
    # 文字色（すべての文字色を取得）
    if (cell.font and cell.font.color and 
        cell.font.color.rgb and 
        isinstance(cell.font.color.rgb, str)):
        styles['color'] = cell.font.color.rgb
    
    return styles

def create_xml_with_styles(wb) -> str:
    """ワークシートの内容とすべてのスタイル情報をXMLに変換"""
    ws = wb.active
    
    # 結合セルの範囲を保存
    merged_cells_ranges = ws.merged_cells.ranges.copy()
    
    # 結合セルを一時的に解除
    for merged_cell_range in merged_cells_ranges:
        ws.unmerge_cells(str(merged_cell_range))
    
    xml_parts = ['<?xml version="1.0" encoding="UTF-8"?>\n<workbook>']
    xml_parts.append('  <worksheet>')
    
    # すべての列幅情報を追加
    for col_letter, col in ws.column_dimensions.items():
        if hasattr(col, 'width') and col.width is not None:
            width = float(col.width)
            xml_parts.append(f'    <column letter="{col_letter}" width="{width:.2f}"/>')
    
    # 結合セルを再設定
    for merged_cell_range in merged_cells_ranges:
        ws.merge_cells(str(merged_cell_range))
    
    # すべての行高さ情報を追加
    for row_idx, row in ws.row_dimensions.items():
        if row.height:
            xml_parts.append(f'    <row index="{row_idx}" height="{row.height:.2f}"/>')
    
    # データ範囲を取得
    data_rows = list(ws.rows)
    if not data_rows:
        return "<workbook></workbook>"
    
    # セルの値をXMLエスケープする関数
    def escape_xml(text):
        if not isinstance(text, str):
            text = str(text)
        return (text.replace('&', '&amp;')
                   .replace('<', '&lt;')
                   .replace('>', '&gt;')
                   .replace('"', '&quot;')
                   .replace("'", '&apos;'))
    
    # すべてのセル情報を処理
    for row in data_rows:
        row_idx = row[0].row
        xml_parts.append(f'    <row index="{row_idx}">')
        
        for cell in row:
            # 値の有無に関わらずすべてのセルを出力
            xml_parts.append(f'      <cell ref="{cell.coordinate}">')
            
            if cell.value is not None:
                value = escape_xml(cell.value)
                xml_parts.append(f'        <value>{value}</value>')
            
            # スタイル情報を追加
            styles = get_cell_styles(cell)
            if styles:
                for style_name, style_value in styles.items():
                    if style_name == 'merge':
                        xml_parts.append('        <merge>')
                        xml_parts.append(f'          <start>{style_value["start"]}</start>')
                        xml_parts.append(f'          <end>{style_value["end"]}</end>')
                        xml_parts.append('        </merge>')
                    elif style_name == 'borders':
                        xml_parts.append('        <borders>')
                        for border_side, border_style in style_value.items():
                            xml_parts.append(f'          <border side="{border_side}" style="{border_style}"/>')
                        xml_parts.append('        </borders>')
                    else:
                        xml_parts.append(f'        <{style_name}>{style_value}</{style_name}>')
            
            xml_parts.append('      </cell>')
        
        xml_parts.append('    </row>')
    
    xml_parts.append('  </worksheet>')
    xml_parts.append('</workbook>')
    return '\n'.join(xml_parts)

def get_url_from_file_data(file_data: Any) -> str:
    """ファイルデータからURLを抽出する"""
    if isinstance(file_data, str):
        # 直接URLが渡された場合
        return file_data
    elif isinstance(file_data, list) and len(file_data) > 0:
        # 配列形式で渡された場合
        first_item = file_data[0]
        if isinstance(first_item, dict) and 'url' in first_item:
            return first_item['url']
    elif isinstance(file_data, dict) and 'url' in file_data:
        # 辞書形式で渡された場合
        return file_data['url']
    elif hasattr(file_data, 'url'):
        # Fileオブジェクトの場合
        return file_data.url
    return None

class DifyPluginExcelToXMLTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage, None, None]:
        
        # ファイルオブジェクトを取得
        file_data = tool_parameters.get("file_url")

        if not file_data:
            yield ToolInvokeMessage(
                type="text",
                message={"text": "ファイルが提供されていません。"}
            )
            return

        try:
            # ファイルのURLを取得
            file_url = get_url_from_file_data(file_data)
            
            if not file_url:
                yield ToolInvokeMessage(
                    type="text",
                    message={"text": "ファイルのURLが見つかりません。"}
                )
                return

            # ファイルをダウンロード
            try:
                response = requests.get(file_url)
                response.raise_for_status()
                file_content = response.content
            except Exception as e:
                yield ToolInvokeMessage(
                    type="text",
                    message={"text": f"ファイルのダウンロードに失敗しました: {str(e)}"}
                )
                return

            # ExcelファイルをOpenpyxlで読み込む
            excel_data = BytesIO(file_content)
            wb = openpyxl.load_workbook(excel_data)
            
            # XMLに変換（罫線情報を含む）
            xml_data = create_xml_with_styles(wb)
            
            # XML形式のテキストを返す
            yield ToolInvokeMessage(
                type="text",
                message={"text": xml_data}
            )

        except Exception as e:
            yield ToolInvokeMessage(
                type="text",
                message={"text": f"エラーが発生しました: {str(e)}"}
            )
            return 