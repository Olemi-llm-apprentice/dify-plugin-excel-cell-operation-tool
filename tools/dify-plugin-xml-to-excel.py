from collections.abc import Generator
from typing import Any
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from io import BytesIO
import xml.etree.ElementTree as ET
import base64

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

def apply_cell_styles(cell, style_elements):
    """セルにスタイルを適用"""
    for style_elem in style_elements:
        if style_elem.tag == 'borders':
            for border in style_elem.findall('border'):
                side = border.get('side')
                style = border.get('style')
                if side and style:
                    border_style = Side(style=style)
                    # 既存の罫線情報を保持しながら更新
                    border_args = {
                        'top': cell.border.top if cell.border else None,
                        'bottom': cell.border.bottom if cell.border else None,
                        'left': cell.border.left if cell.border else None,
                        'right': cell.border.right if cell.border else None
                    }
                    border_args[side] = border_style
                    cell.border = Border(**border_args)
        
        elif style_elem.tag == 'merge':
            # mergeタグの中身を直接取得
            merge_info = style_elem.attrib
            if 'start' in merge_info and 'end' in merge_info:
                start = merge_info['start']
                end = merge_info['end']
                try:
                    cell.parent.merge_cells(f'{start}:{end}')
                except ValueError:
                    # 既に結合されているセルの場合はスキップ
                    pass
        
        elif style_elem.tag == 'background':
            from openpyxl.styles import PatternFill
            if style_elem.text:  # テキストが存在する場合のみ適用
                cell.fill = PatternFill(start_color=style_elem.text, 
                                      end_color=style_elem.text, 
                                      fill_type='solid')
        
        elif style_elem.tag == 'color':
            from openpyxl.styles import Font
            if style_elem.text:  # テキストが存在する場合のみ適用
                current_font = cell.font or Font()
                cell.font = Font(color=style_elem.text,
                               name=current_font.name,
                               size=current_font.size,
                               bold=current_font.bold,
                               italic=current_font.italic)

def create_excel_from_xml(xml_text: str) -> BytesIO:
    """XML形式のテキストからExcelファイルを生成"""
    # 新しいワークブックを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    
    try:
        # XMLをパース
        root = ET.fromstring(xml_text)
        worksheet = root.find('worksheet')
        if worksheet is None:
            raise ValueError("worksheetタグが見つかりません")
        
        # 各行を処理
        for row_elem in worksheet.findall('row'):
            try:
                row_idx = int(row_elem.get('index'))
            except (ValueError, TypeError):
                continue
            
            # 各セルを処理
            for cell_elem in row_elem.findall('cell'):
                ref = cell_elem.get('ref')
                if not ref:
                    continue
                
                try:
                    # セルの値を設定
                    value_elem = cell_elem.find('value')
                    if value_elem is not None and value_elem.text is not None:
                        ws[ref] = value_elem.text
                    
                    # スタイルを適用
                    apply_cell_styles(ws[ref], cell_elem)
                except Exception as e:
                    print(f"セル {ref} の処理中にエラー: {str(e)}")
                    continue
        
        # メモリ上にExcelファイルを保存
        excel_io = BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        return excel_io
        
    except Exception as e:
        raise ValueError(f"XMLの処理中にエラー: {str(e)}")

class DifyPluginXMLToExcelTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage, None, None]:
        try:
            # XMLテキストを取得
            xml_text = tool_parameters.get("xml_text")
            if not xml_text:
                yield ToolInvokeMessage(
                    type="text",
                    message={"text": "XMLテキストが提供されていません。"}
                )
                return

            # XMLからExcelファイルを生成
            excel_io = create_excel_from_xml(xml_text)
            excel_data = excel_io.getvalue()
            
            # blobメッセージとしてファイルを返す
            yield self.create_blob_message(
                blob=excel_data,
                meta={
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "filename": "converted.xlsx"
                }
            )

        except Exception as e:
            yield ToolInvokeMessage(
                type="text",
                message={"text": f"エラーが発生しました: {str(e)}"}
            )
            return 