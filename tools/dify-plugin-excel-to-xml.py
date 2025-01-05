from collections.abc import Generator
from typing import Any
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

def get_important_cell_styles(cell) -> dict:
    """デフォルトから変更のある重要なスタイル情報のみを取得"""
    styles = {}
    
    # セル結合情報の取得
    if cell.parent.merged_cells:  # merged_cellsはワークシートの属性
        for merged_range in cell.parent.merged_cells.ranges:
            if cell.coordinate in merged_range:
                styles['merge'] = {
                    'start': merged_range.coord.split(':')[0],  # 結合開始セル
                    'end': merged_range.coord.split(':')[1]     # 結合終了セル
                }
                break
    
    # 罫線情報（デフォルトのNoneまたはthinは除外）
    borders = {}
    for side in ['top', 'bottom', 'left', 'right']:
        border = getattr(cell.border, side)
        if border.style and border.style not in ['thin', None]:
            borders[side] = border.style
    if borders:
        styles['borders'] = borders
    
    # 背景色（デフォルトの白や透明以外で、かつ有効な値の場合のみ）
    if (cell.fill and cell.fill.start_color and 
        cell.fill.start_color.rgb and 
        cell.fill.start_color.rgb not in ['FFFFFFFF', '00000000'] and  # 白と透明を除外
        isinstance(cell.fill.start_color.rgb, str)):
        styles['background'] = cell.fill.start_color.rgb
    
    # 文字色（デフォルトの黒以外で、かつ有効な値の場合のみ）
    if (cell.font and cell.font.color and 
        cell.font.color.rgb and 
        cell.font.color.rgb != 'FF000000' and
        isinstance(cell.font.color.rgb, str)):
        styles['color'] = cell.font.color.rgb
    
    return styles if styles else None

def create_xml_with_styles(wb) -> str:
    """ワークシートの内容と変更のあるスタイル情報のみをXMLに変換"""
    ws = wb.active
    xml_parts = ['<?xml version="1.0" encoding="UTF-8"?>\n<workbook>']
    
    # データ範囲を取得
    data_rows = list(ws.rows)
    if not data_rows:
        return "<workbook></workbook>"
    
    xml_parts.append('  <worksheet>')
    
    for row_idx, row in enumerate(data_rows, 1):
        has_content = False
        row_parts = []
        
        for col_idx, cell in enumerate(row, 1):
            col_letter = get_column_letter(col_idx)
            value = cell.value if cell.value is not None else ""
            styles = get_important_cell_styles(cell)
            
            # 値かスタイルがある場合のみ出力
            if value or styles:
                has_content = True
                cell_parts = [f'      <cell ref="{col_letter}{row_idx}">']
                
                if value:
                    cell_parts.append(f'        <value>{value}</value>')
                
                if styles:
                    for style_type, style_value in styles.items():
                        if isinstance(style_value, dict):
                            # 罫線情報の場合
                            if style_value:  # 空の辞書は出力しない
                                cell_parts.append(f'        <{style_type}>')
                                for side, style in style_value.items():
                                    cell_parts.append(f'          <border side="{side}" style="{style}"/>')
                                cell_parts.append(f'        </{style_type}>')
                        else:
                            # 背景色や文字色の場合（有効な値の場合のみ）
                            cell_parts.append(f'        <{style_type}>{style_value}</{style_type}>')
                
                cell_parts.append('      </cell>')
                row_parts.extend(cell_parts)
        
        # 行に内容がある場合のみ出力
        if has_content:
            xml_parts.append(f'    <row index="{row_idx}">')
            xml_parts.extend(row_parts)
            xml_parts.append('    </row>')
    
    xml_parts.append('  </worksheet>')
    xml_parts.append('</workbook>')
    
    return '\n'.join(xml_parts)

def get_url_from_file_data(file_data: Any) -> str:
    """ファイルデータからURLを抽出する"""
    if isinstance(file_data, str):
        # 直接URLが渡された場合
        return file_data
    elif isinstance(file_data, list) and len(file_data) > 0:
        # 配列形式で渡された場合
        first_item = file_data[0]
        if isinstance(first_item, dict) and 'url' in first_item:
            return first_item['url']
    elif isinstance(file_data, dict) and 'url' in file_data:
        # 辞書形式で渡された場合
        return file_data['url']
    elif hasattr(file_data, 'url'):
        # Fileオブジェクトの場合
        return file_data.url
    return None

class DifyPluginExcelToXMLTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage, None, None]:
        # フバッグ用にパラメータの内容を出力（ファイルオブジェクトを除外）
        try:
            print("Debug - tool_parameters keys:", tool_parameters.keys())
        except:
            print("Debug - tool_parameters: <not serializable>")
        
        # ファイルオブジェクトを取得
        file_data = tool_parameters.get("file_url")
        try:
            if isinstance(file_data, dict):
                print("Debug - file_data keys:", file_data.keys())
            elif isinstance(file_data, list):
                print("Debug - file_data is list of length:", len(file_data))
                if len(file_data) > 0:
                    print("Debug - first item keys:", file_data[0].keys())
            else:
                print("Debug - file_data type:", type(file_data))
        except:
            print("Debug - file_data: <not serializable>")

        if not file_data:
            yield ToolInvokeMessage(
                type="text",
                message={"text": "ファイルが提供されていません。"}
            )
            return

        try:
            # ファイルのURLを取得
            file_url = get_url_from_file_data(file_data)
            print("Debug - file_url:", file_url)  # 取得したURLを確認
            
            if not file_url:
                yield ToolInvokeMessage(
                    type="text",
                    message={"text": "ファイルのURLが見つかりません。"}
                )
                return

            # ファイルをダウンロード
            try:
                response = requests.get(file_url)
                response.raise_for_status()
                file_content = response.content
            except Exception as e:
                yield ToolInvokeMessage(
                    type="text",
                    message={"text": f"ファイルのダウンロードに失敗しました: {str(e)}"}
                )
                return

            # ExcelファイルをOpenpyxlで読み込む
            excel_data = BytesIO(file_content)
            wb = openpyxl.load_workbook(excel_data)
            
            # XMLに変換（罫線情報を含む）
            xml_data = create_xml_with_styles(wb)
            
            # XML形式のテキストを返す
            yield ToolInvokeMessage(
                type="text",
                message={"text": xml_data}
            )

        except Exception as e:
            yield ToolInvokeMessage(
                type="text",
                message={"text": f"エラーが発生しました: {str(e)}"}
            )
            return 